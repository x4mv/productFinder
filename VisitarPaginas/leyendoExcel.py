import os
import openpyxl
from tkinter import filedialog

#Arreglo con los codigos 
arrayCodigos = []

#Arreglo con los Nombres 
arrayNombreProducto = []

def seleccionar_archivo():
            archivo = filedialog.askopenfilename(
            title="Seleccionar archivo", 
            filetypes=[("Archivos Python", "*"), ("Todos los archivos", "*.*")]
        )
            if archivo:
                # El usuario ha seleccionado un archivo, puedes hacer algo con él
                return archivo


def leer():
        
    
    ruta = seleccionar_archivo()

    book = openpyxl.load_workbook(ruta)

    sheet = book.active

    # Itera a través de todas las filas en la hoja de trabajo
    for filaCodigo in sheet.iter_rows(min_row=4, max_row = sheet.max_row, min_col =2, max_col = 3, values_only=True):
        # `fila` es una tupla que contiene los valores de cada celda en la fila actual
        if filaCodigo[0] != None:
            arrayCodigos.append(filaCodigo[0])
            arrayNombreProducto.append(filaCodigo[1])
